"""Génération d'exports Excel soignés (en-tête aux couleurs CIE, colonnes
dimensionnées, ligne d'en-tête figée et filtrable, titre optionnel), réutilisé par
tous les exports de l'application (clients, prospection, réseau, messagerie) pour un
rendu propre et homogène plutôt que le tableau brut de pandas."""

import io

from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CIE_ORANGE = "F7941E"
CIE_VERT = "2E9E4F"
GRIS_LEGER = "F2F2F2"

CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def excel_response(df, filename, sheet_name="Export", titre=None, couleur_entete=CIE_ORANGE):
    """Retourne une HttpResponse .xlsx stylée à partir d'un DataFrame pandas.

    - titre : ligne de titre en gras au-dessus du tableau (optionnel) ;
    - en-tête : fond couleur CIE + texte blanc gras, centré, renvoi à la ligne ;
    - colonnes dimensionnées d'après le contenu (échantillon), largeur bornée ;
    - ligne d'en-tête figée + filtre automatique.
    Le style ne s'applique qu'à l'en-tête et aux dimensions (rapide même sur de
    gros exports) ; les lignes de données restent non décorées pour la performance."""
    import pandas as pd  # import local (pandas lourd, pas nécessaire au chargement du module)

    buffer = io.BytesIO()
    startrow = 2 if titre else 0
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow)
        ws = writer.sheets[sheet_name]

        if titre:
            cell = ws.cell(row=1, column=1, value=titre)
            cell.font = Font(size=14, bold=True, color=couleur_entete)

        # Si le résultat filtré est vide, pandas produit un DataFrame SANS
        # colonnes (pas seulement sans lignes) : get_column_letter(0) plus bas
        # lèverait "Invalid column index 0". On saute la mise en forme de
        # l'en-tête dans ce cas plutôt que de planter (un export vide reste un
        # export valide, juste sans style).
        if len(df.columns) > 0:
            header_row = startrow + 1
            fill = PatternFill("solid", fgColor=couleur_entete)
            font = Font(bold=True, color="FFFFFF")
            thin = Side(style="thin", color="D9D9D9")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col_idx, nom in enumerate(df.columns, start=1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                echantillon = df.iloc[:, col_idx - 1].head(300)
                largeur = max([len(str(nom))] + [len(str(v)) for v in echantillon]) + 2
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(largeur, 12), 48)

            ws.row_dimensions[header_row].height = 28
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
            derniere_col = get_column_letter(len(df.columns))
            ws.auto_filter.ref = f"A{header_row}:{derniere_col}{header_row + len(df)}"

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type=CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

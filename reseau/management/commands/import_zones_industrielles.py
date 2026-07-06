"""Importe le référentiel des Zones Industrielles et de leurs départs HTA.

Source : informations clients/dcb/Support Technique/ZI_Departures_SS.xlsx, une
seule feuille organisée en blocs (pas un tableau plat) : une ligne d'en-tête de
zone ("ZI de YOPOUGON" ou, cas particulier, "Poste source de VRIDI"), suivie
d'une ligne de sous-en-tête ("N°"/"PS"/"Départ"), puis des lignes de données
jusqu'à la prochaine zone. C'est ce référentiel qui permet de classer un
incident/des travaux par zone (cf. reseau.models.DepartZoneIndustrielle),
puisque ni INCIBCC ni MANTBCC ne portent eux-mêmes cette information.

Toujours rejoué en entier (delete + recreate) plutôt que complété : ce fichier
est un référentiel statique (la liste des départs qui alimentent chaque zone),
pas un historique cumulatif comme INCIBCC/MANTBCC.
"""

import openpyxl
from django.core.management.base import BaseCommand

from importers.utils import INFO_CLIENTS_DIR
from reseau.models import DepartZoneIndustrielle, ZoneIndustrielle

PREFIXES_ZONE = ("ZI de ", "Poste source de ")


def _nom_zone(valeur):
    """"ZI de YOPOUGON" -> "YOPOUGON", "Poste source de VRIDI" -> "VRIDI"."""
    texte = str(valeur).strip()
    for prefixe in PREFIXES_ZONE:
        if texte.startswith(prefixe):
            return texte[len(prefixe):].strip()
    return None


class Command(BaseCommand):
    help = "Importe ZoneIndustrielle et DepartZoneIndustrielle depuis ZI_Departures_SS.xlsx."

    def handle(self, *args, **options):
        path = INFO_CLIENTS_DIR / "dcb" / "Support Technique" / "ZI_Departures_SS.xlsx"
        self.stdout.write(f"Lecture de {path}...")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        DepartZoneIndustrielle.objects.all().delete()
        ZoneIndustrielle.objects.all().delete()

        zone_courante = None
        nb_departs = 0
        for row in ws.iter_rows(min_row=1, values_only=True):
            premiere_cellule = row[0]
            if premiere_cellule is None:
                continue
            nom_zone = _nom_zone(premiere_cellule)
            if nom_zone:
                zone_courante, _ = ZoneIndustrielle.objects.get_or_create(nom=nom_zone)
                continue
            # Ligne de sous-en-tête ("N°"/"PS"/"Départ", abîmée en mojibake selon
            # l'encodage de lecture) : ni une zone, ni une ligne de données utile.
            if not isinstance(premiere_cellule, (int, float)):
                continue
            if zone_courante is None or row[2] is None:
                continue
            DepartZoneIndustrielle.objects.get_or_create(
                zone=zone_courante,
                nom_depart=str(row[2]).strip(),
                defaults={"poste_source": str(row[1] or "").strip()},
            )
            nb_departs += 1

        wb.close()
        nb_zones = ZoneIndustrielle.objects.count()
        self.stdout.write(self.style.SUCCESS(f"{nb_zones} zones industrielles, {nb_departs} départs rattachés."))
